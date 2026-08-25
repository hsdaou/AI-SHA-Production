"""
School Timetable Search Interface
Connects: Student_Database_1.xlsx, Teacher_Schedule_Database.xlsx, school_timetable.db
Allows searching for free students at specific periods/days per grade.
A student is free if their section has a free period OR if the student
is not enrolled in the subject being taught at that period.
"""

import sqlite3
import os
import re
from flask import Flask, render_template, request, jsonify
import openpyxl
from collections import defaultdict

app = Flask(__name__)

# Robot-facing surface (AI-SHA). Separate blueprint so the timetable logic below
# stays untouched: /api/robot/* is authenticated and never returns a list of
# named students to the caller. See robot_api.py.
from robot_api import robot_api as _robot_bp  # noqa: E402
app.register_blueprint(_robot_bp)

# ── File paths ─────────────────────────────────────────────────
# Configurable so the app runs on the machine it is deployed to. The previous
# hard-coded C:\Users\hsdao\Downloads meant it only ever ran on one laptop —
# and the AI-SHA integration needs it reachable from the robot.
BASE = os.environ.get("SCHOOL_DATA_DIR") or os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("SCHOOL_DB_PATH") or os.path.join(BASE, "school_timetable.db")
STUDENT_XLSX = os.environ.get("SCHOOL_STUDENT_XLSX") or os.path.join(BASE, "Student_Database_1.xlsx")
TEACHER_XLSX = os.environ.get("SCHOOL_TEACHER_XLSX") or os.path.join(BASE, "Teacher_Schedule_Database.xlsx")


# ── Subject matching ──────────────────────────────────────────
# The timetable uses short/generic subject names like "Mathematics",
# "Physics", "Arabic Language". The student enrollment DB uses detailed
# names like "Mathematics N2", "Physics N3", "Arabic Language- Arabs".
# We need fuzzy matching to determine if a student is enrolled in the
# subject being taught.

def normalize_subject(name):
    """Normalize a subject name for matching: strip levels, suffixes, and expand abbreviations."""
    if not name:
        return ""
    n = name.strip().lower()
    # Remove level/track suffixes:  N1, N2, N3, N4, L, L1, M, M1, MA, TN, TN1, etc.
    #
    # The level code must be a SEPARATE, SHORT token - hence the required
    # separator and the bounded repeats. The previous pattern was
    #     \s*[-]?\s*(n[0-9]*|l[0-9]*|m[0-9a-z]*|tn[0-9]*)$
    # where the separator was optional and m[0-9a-z]* was unbounded, so it ate
    # whole words off the end of ordinary subject names:
    #     "Mathematics" -> ""        (matched nobody: every student looked FREE)
    #     "Music"       -> ""        (same)
    #     "Computing"   -> "co"      (matched far too much: false BUSY)
    #     "Latin"       -> "lati", "Nutrition" -> "nutritio"
    # Two subjects normalising to empty covered 395 of 3294 timetable rows, and
    # because the same function normalises the ENROLMENT side, students enrolled
    # in "Mathematics" were invisible too. A grade-8 period-3 query returned 210
    # of 269 students "free" when the whole grade was in class.
    n = re.sub(r"(?:\s+|\s*-\s*)(?:n\d{0,2}|l\d{0,2}|m[a-z]?\d{0,2}|tn\d{0,2})$", "", n)
    # Remove cultural descriptors: "- Arabs", "- Non Arabs", "Non A"
    n = re.sub(r"\s*[-]\s*(arabs?|non arabs?|non a)$", "", n)
    # Expand common short forms
    n = n.replace("math ", "mathematics ")
    if n == "math":
        n = "mathematics"
    # Strip trailing punctuation/whitespace
    n = n.strip().rstrip("-").strip()
    return n


def student_is_enrolled(student_subjects, timetable_subject):
    """
    Check if a student (given their enrolled subject names) is enrolled
    in the timetable subject being taught.

    Returns True if the student has a matching enrollment.
    """
    if not timetable_subject:
        return False

    tt_norm = normalize_subject(timetable_subject)
    if not tt_norm:
        return False

    for enrolled_subj in student_subjects:
        en_norm = normalize_subject(enrolled_subj)
        if not en_norm:
            continue
        # Match if either is a prefix of the other
        if en_norm.startswith(tt_norm) or tt_norm.startswith(en_norm):
            return True
        # Also match exact equality after normalization
        if en_norm == tt_norm:
            return True

    return False


# ── Section classification ────────────────────────────────────
# Some timetable sections have no students in the Student DB (e.g. if
# the grade has students in A-H but the timetable also has I, J).
# These are "level sections" — shown in the UI but with 0 students.

def get_level_sections(grade):
    """
    Return the set of timetable section letters for this grade
    that have NO students directly assigned in the Student DB.
    """
    cache_key = f"level_sections_{grade}"
    if cache_key in _cache:
        return _cache[cache_key]

    students = load_students()
    grade_str = str(grade).zfill(2)
    student_sections = set(s["section"] for s in students if s["grade"] == grade_str)

    # If no students exist for this grade, there are no level sections
    if not student_sections:
        _cache[cache_key] = set()
        return set()

    tt_sections = get_sections_for_grade(grade)

    level_secs = set()
    for tt_sec in tt_sections:
        if tt_sec not in student_sections:
            level_secs.add(tt_sec)

    _cache[cache_key] = level_secs
    return level_secs


# ── Data loaders (cached) ─────────────────────────────────────
_cache = {}

def get_db():
    """Get SQLite connection."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def load_students():
    """Load student master list from Excel. Returns list of dicts."""
    if "students" in _cache:
        return _cache["students"]

    wb = openpyxl.load_workbook(STUDENT_XLSX, read_only=True)
    ws = wb["Student Master List"]
    students = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(row)]
            continue
        rec = dict(zip(headers, row))
        students.append({
            "computer_number": str(rec.get("Computer Number", "")),
            "first_name": str(rec.get("First Name", "")),
            "last_name": str(rec.get("Last Name", "")),
            "grade": str(rec.get("Class", "")).zfill(2),
            "section": str(rec.get("Section", "")),
            "email": str(rec.get("Email", "")),
            "num_subjects": rec.get("Number of Subjects", 0),
        })
    wb.close()
    _cache["students"] = students
    return students


def load_student_enrollments():
    """
    Load subject enrollment. Returns two structures:
    1. enrollments: (grade, section, computer_number) -> [subject dicts]
    2. student_subjects: computer_number -> set of subject names
    """
    if "enrollments" in _cache:
        return _cache["enrollments"], _cache["student_subjects"]

    wb = openpyxl.load_workbook(STUDENT_XLSX, read_only=True)
    ws = wb["Subject Enrollment"]
    enrollments = defaultdict(list)
    student_subjects = defaultdict(set)
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(row)]
            continue
        rec = dict(zip(headers, row))
        grade = str(rec.get("Class", "")).zfill(2)
        section = str(rec.get("Section", ""))
        comp_num = str(rec.get("Computer Number", ""))
        subject = str(rec.get("Subject Name", ""))
        subject_code = str(rec.get("Subject Code", ""))
        teacher = str(rec.get("Teacher", ""))

        key = (grade, section, comp_num)
        enrollments[key].append({
            "subject_code": subject_code,
            "subject_name": subject,
            "teacher": teacher,
            "course_group": str(rec.get("Course Group", "")),
        })
        if subject:
            student_subjects[comp_num].add(subject)

    wb.close()
    _cache["enrollments"] = enrollments
    _cache["student_subjects"] = student_subjects
    return enrollments, student_subjects


def load_teacher_directory():
    """Load teacher directory from Teacher Schedule Excel."""
    if "teacher_dir" in _cache:
        return _cache["teacher_dir"]

    wb = openpyxl.load_workbook(TEACHER_XLSX, read_only=True)
    ws = wb["Teacher Directory"]
    teachers = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(row)]
            continue
        rec = dict(zip(headers, row))
        teachers.append({
            "code": str(rec.get("Teacher Code", "")),
            "title": str(rec.get("Title", "")),
            "name": str(rec.get("Teacher Name", "")),
            "total_periods": rec.get("Total Periods/Week", 0),
            "subjects": str(rec.get("Subjects Taught", "")),
            "rooms": str(rec.get("Rooms/Classes", "")),
        })
    wb.close()
    _cache["teacher_dir"] = teachers
    return teachers


def load_teacher_full_schedule():
    """Load full teacher schedule. Returns dict: (teacher_code, day, period) -> schedule info."""
    if "teacher_sched" in _cache:
        return _cache["teacher_sched"]

    wb = openpyxl.load_workbook(TEACHER_XLSX, read_only=True)
    ws = wb["Full Schedule"]
    schedule = {}
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(row)]
            continue
        rec = dict(zip(headers, row))
        tc = str(rec.get("Teacher Code", ""))
        day = str(rec.get("Day", ""))
        period = str(rec.get("Period", ""))
        key = (tc, day, period)
        schedule[key] = {
            "teacher_code": tc,
            "teacher_name": str(rec.get("Teacher Name", "")),
            "day": day,
            "period": period,
            "time": str(rec.get("Time", "")),
            "subject": str(rec.get("Subject", "")),
            "group_code": str(rec.get("Group Code", "")),
            "room": str(rec.get("Room/Class", "")),
        }
    wb.close()
    _cache["teacher_sched"] = schedule
    return schedule


def get_teacher_lookup():
    """Build a teacher_code → full name mapping from the Teacher Directory."""
    if "teacher_lookup" in _cache:
        return _cache["teacher_lookup"]

    teachers = load_teacher_directory()
    lookup = {}
    for t in teachers:
        code = t["code"]
        title = t.get("title", "").strip()
        name = t.get("name", "").strip()
        lookup[code] = f"{title} {name}".strip() if name else code

    _cache["teacher_lookup"] = lookup
    return lookup


def resolve_teacher(teacher_code):
    """Resolve a teacher code to their full name. Falls back to the code itself."""
    if not teacher_code:
        return None
    lookup = get_teacher_lookup()
    return lookup.get(teacher_code, teacher_code)


def get_class_timetable(grade, section, day):
    """
    Get timetable for a specific class-section and day from the SQLite DB.
    Returns list of period entries.
    """
    section_code = f"{int(grade):02d}-{section}"
    conn = get_db()
    rows = conn.execute("""
        SELECT te.period_name, te.period_time, te.subject, te.teacher_code,
               te.group_code, te.is_free, te.raw_cell
        FROM timetable_entries te
        JOIN class_sections cs ON te.class_section_id = cs.id
        WHERE cs.section_code = ? AND te.day = ?
        ORDER BY
            CASE
                WHEN te.period_name LIKE 'Period %' THEN CAST(REPLACE(te.period_name, 'Period ', '') AS INTEGER)
                WHEN te.period_name = 'Lunch' THEN 50
                WHEN te.period_name LIKE 'SLO%' THEN 60
                WHEN te.period_name = 'After School' THEN 100
                ELSE 90
            END
    """, (section_code, day)).fetchall()
    conn.close()

    result = []
    for r in rows:
        entry = dict(r)
        # Free status comes directly from the database
        entry["effectively_free"] = entry["is_free"] == 1
        # Resolve teacher code to name
        entry["teacher_name"] = resolve_teacher(entry["teacher_code"])
        result.append(entry)
    return result


def get_all_grades():
    """Get all unique grades from the timetable DB."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT grade FROM class_sections ORDER BY grade"
    ).fetchall()
    conn.close()
    return [r["grade"] for r in rows]


def get_sections_for_grade(grade):
    """Get all sections for a given grade."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT section_letter FROM class_sections WHERE grade = ? ORDER BY section_letter",
        (grade,)
    ).fetchall()
    conn.close()
    return [r["section_letter"] for r in rows]


def get_periods_for_grade(grade):
    """Get all period names for a grade from the timetable."""
    conn = get_db()
    rows = conn.execute("""
        SELECT DISTINCT te.period_name
        FROM timetable_entries te
        JOIN class_sections cs ON te.class_section_id = cs.id
        WHERE cs.grade = ?
        ORDER BY
            CASE
                WHEN te.period_name LIKE 'Period %' THEN CAST(REPLACE(te.period_name, 'Period ', '') AS INTEGER)
                WHEN te.period_name = 'Lunch' THEN 50
                WHEN te.period_name LIKE 'SLO%' THEN 60
                WHEN te.period_name = 'After School' THEN 100
                ELSE 90
            END
    """, (grade,)).fetchall()
    conn.close()
    return [r["period_name"] for r in rows]


def find_free_students(grade, day, period_name):
    """
    Find students who are FREE at a given grade/day/period.

    A student is free if:
    1. Their section has a free period (no class scheduled), OR
    2. Their section has a class but the student is NOT enrolled in that subject.

    Returns (free_students, busy_students) lists.
    """
    students = load_students()
    level_secs = get_level_sections(grade)
    _, student_subjects = load_student_enrollments()

    grade_str = str(grade).zfill(2)
    grade_students = [s for s in students if s["grade"] == grade_str]

    # Group students by their section
    students_by_section = defaultdict(list)
    for s in grade_students:
        students_by_section[s["section"]].append(s)

    free_students = []
    busy_students = []

    for section, section_students in students_by_section.items():
        if section in level_secs:
            continue  # no students in the DB for this section

        timetable = get_class_timetable(grade, section, day)
        period_entry = None
        for entry in timetable:
            if entry["period_name"] == period_name:
                period_entry = entry
                break

        teacher_display = None
        tt_subject = None
        if period_entry:
            teacher_display = period_entry.get("teacher_name") or resolve_teacher(period_entry["teacher_code"])
            tt_subject = period_entry["subject"]

        for s in section_students:
            cn = s["computer_number"]
            enrolled_subjects = student_subjects.get(cn, set())

            if period_entry is None:
                # No timetable entry → free
                free_students.append({
                    **s,
                    "status": "free",
                    "reason": "No timetable entry",
                    "subject": None,
                    "teacher": None,
                })
            elif period_entry["effectively_free"] or not tt_subject:
                # Section is free or no subject assigned
                free_students.append({
                    **s,
                    "status": "free",
                    "reason": "Free period",
                    "subject": tt_subject,
                    "teacher": teacher_display,
                })
            elif not student_is_enrolled(enrolled_subjects, tt_subject):
                # Section has class but student is NOT enrolled in this subject
                free_students.append({
                    **s,
                    "status": "free",
                    "reason": f"Not enrolled in: {tt_subject}",
                    "subject": tt_subject,
                    "teacher": teacher_display,
                })
            else:
                # Student is enrolled → busy
                busy_students.append({
                    **s,
                    "status": "busy",
                    "reason": f"In class: {tt_subject}",
                    "subject": tt_subject,
                    "teacher": teacher_display,
                })

    return free_students, busy_students


def find_free_teachers(day, period_num):
    """Find teachers who are free at a given day/period."""
    teacher_sched = load_teacher_full_schedule()
    teacher_dir = load_teacher_directory()

    # Build set of busy teachers
    busy_teachers = set()
    for (tc, d, p), info in teacher_sched.items():
        if d == day and str(p) == str(period_num):
            busy_teachers.add(tc)

    free = []
    busy = []
    for t in teacher_dir:
        if t["code"] in busy_teachers:
            # Get what they're teaching
            key = (t["code"], day, str(period_num))
            sched = teacher_sched.get(key, {})
            busy.append({
                **t,
                "status": "busy",
                "current_subject": sched.get("subject", ""),
                "current_room": sched.get("room", ""),
            })
        else:
            free.append({
                **t,
                "status": "free",
                "current_subject": None,
                "current_room": None,
            })

    return free, busy


# ── Flask Routes ───────────────────────────────────────────────

@app.route("/")
def index():
    grades = get_all_grades()
    return render_template("index.html", grades=grades)


@app.route("/api/sections/<int:grade>")
def api_sections(grade):
    sections = get_sections_for_grade(grade)
    return jsonify(sections)


@app.route("/api/periods/<int:grade>")
def api_periods(grade):
    periods = get_periods_for_grade(grade)
    return jsonify(periods)


@app.route("/api/free-students")
def api_free_students():
    grade = request.args.get("grade", type=int)
    day = request.args.get("day", "Monday")
    period = request.args.get("period", "Period 1")

    if not grade:
        return jsonify({"error": "Grade is required"}), 400

    free, busy = find_free_students(grade, day, period)

    # Group free students by section
    by_section = defaultdict(list)
    for s in free:
        by_section[s["section"]].append(s)

    busy_by_section = defaultdict(list)
    for s in busy:
        busy_by_section[s["section"]].append(s)

    return jsonify({
        "grade": grade,
        "day": day,
        "period": period,
        "free_count": len(free),
        "busy_count": len(busy),
        "total_count": len(free) + len(busy),
        "free_students": free,
        "busy_students": busy,
        "free_by_section": dict(by_section),
        "busy_by_section": dict(busy_by_section),
    })


@app.route("/api/free-teachers")
def api_free_teachers():
    day = request.args.get("day", "Monday")
    period = request.args.get("period", "1")

    free, busy = find_free_teachers(day, period)

    return jsonify({
        "day": day,
        "period": period,
        "free_count": len(free),
        "busy_count": len(busy),
        "free_teachers": free,
        "busy_teachers": busy,
    })


@app.route("/api/class-timetable")
def api_class_timetable():
    grade = request.args.get("grade", type=int)
    section = request.args.get("section", "A")
    day = request.args.get("day", "Monday")

    if not grade:
        return jsonify({"error": "Grade is required"}), 400

    timetable = get_class_timetable(grade, section, day)
    return jsonify({
        "grade": grade,
        "section": section,
        "day": day,
        "periods": [dict(t) for t in timetable],
    })


@app.route("/api/section-status")
def api_section_status():
    """
    Get free/busy status for ALL timetable sections in a grade for a specific
    day/period. Shows every section from the timetable (including level sections
    like 10-L, 11-LA) with student counts.

    Level sections are shown for reference only — they don't affect student
    free/busy counts. Only real sections (with students) contribute to totals.
    """
    grade = request.args.get("grade", type=int)
    day = request.args.get("day", "Monday")
    period = request.args.get("period", "Period 1")

    if not grade:
        return jsonify({"error": "Grade required"}), 400

    tt_sections = get_sections_for_grade(grade)
    students = load_students()
    grade_str = str(grade).zfill(2)
    level_secs = get_level_sections(grade)

    # Collect actual student sections and their counts
    student_section_counts = defaultdict(int)
    for s in students:
        if s["grade"] == grade_str:
            student_section_counts[s["section"]] += 1

    result = []

    for tt_sec in tt_sections:
        tt = get_class_timetable(grade, tt_sec, day)
        entry = None
        for t in tt:
            if t["period_name"] == period:
                entry = t
                break

        is_level = tt_sec in level_secs
        count = student_section_counts.get(tt_sec, 0)

        if entry is None:
            result.append({
                "section": tt_sec,
                "status": "free",
                "subject": None,
                "teacher": None,
                "student_count": count,
                "reason": "No class scheduled",
                "is_level_section": is_level,
            })
        elif entry["effectively_free"]:
            result.append({
                "section": tt_sec,
                "status": "free",
                "subject": entry["subject"],
                "teacher": entry.get("teacher_name") or resolve_teacher(entry["teacher_code"]),
                "student_count": count,
                "reason": "Free period",
                "is_level_section": is_level,
            })
        else:
            result.append({
                "section": tt_sec,
                "status": "busy",
                "subject": entry["subject"],
                "teacher": entry.get("teacher_name") or resolve_teacher(entry["teacher_code"]),
                "student_count": count,
                "reason": entry["subject"],
                "is_level_section": is_level,
            })

    return jsonify({
        "grade": grade,
        "day": day,
        "period": period,
        "sections": result,
        "total_free": sum(s["student_count"] for s in result if s["status"] == "free" and not s["is_level_section"]),
        "total_busy": sum(s["student_count"] for s in result if s["status"] == "busy" and not s["is_level_section"]),
    })


@app.route("/api/free-students-day")
def api_free_students_day():
    """
    Find students who have at least one free period on a given day.
    A student is free at a period if:
    1. Their section has a free period, OR
    2. Their section has a class but the student is NOT enrolled in that subject.
    Returns each student with a list of which periods they are free.
    """
    grade = request.args.get("grade", type=int)
    day = request.args.get("day", "Monday")

    if not grade:
        return jsonify({"error": "Grade is required"}), 400

    periods = get_periods_for_grade(grade)
    students = load_students()
    level_secs = get_level_sections(grade)
    _, student_subjects = load_student_enrollments()

    grade_str = str(grade).zfill(2)
    grade_students = [s for s in students if s["grade"] == grade_str]

    # Group students by their section
    students_by_section = defaultdict(list)
    for s in grade_students:
        students_by_section[s["section"]].append(s)

    # Build per-section timetable for the day
    section_timetables = {}
    for section in students_by_section:
        if section in level_secs:
            continue
        section_timetables[section] = get_class_timetable(grade, section, day)

    # For each student, determine which periods they are free
    student_free_periods = {}

    for section, section_students in students_by_section.items():
        if section in level_secs:
            continue

        timetable = section_timetables.get(section, [])
        period_lookup = {}
        for entry in timetable:
            period_lookup[entry["period_name"]] = entry

        for s in section_students:
            cn = s["computer_number"]
            enrolled_subjects = student_subjects.get(cn, set())
            free_periods = []

            for period in periods:
                entry = period_lookup.get(period)
                if entry is None or entry["effectively_free"] or not entry["subject"]:
                    # Section has no class, is free, or no subject assigned
                    free_periods.append(period)
                elif not student_is_enrolled(enrolled_subjects, entry["subject"]):
                    # Section has a class but student isn't enrolled
                    free_periods.append(period)

            if free_periods:
                student_free_periods[cn] = {
                    **s,
                    "free_periods": free_periods,
                    "free_count": len(free_periods),
                    "total_periods": len(periods),
                }

    # Sort by number of free periods (most free first)
    result = sorted(student_free_periods.values(), key=lambda x: -x["free_count"])

    return jsonify({
        "grade": grade,
        "day": day,
        "periods": periods,
        "total_students": len(grade_students),
        "students_with_free": len(result),
        "students": result,
    })


if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("  School Timetable Search Interface")
    print("  International School of Choueifat, Sharjah")
    print("=" * 60)
    print(f"  DB:       {DB_PATH}")
    print(f"  Students: {STUDENT_XLSX}")
    print(f"  Teachers: {TEACHER_XLSX}")
    print("=" * 60)
    host = os.environ.get("SCHOOL_HOST", "127.0.0.1")
    port = int(os.environ.get("SCHOOL_PORT", "5000"))
    # debug=True exposes the Werkzeug console, which allows ARBITRARY CODE
    # EXECUTION to anyone who can reach the port. That was survivable while this
    # was localhost-only; it is not once the robot has to reach it. Opt in
    # explicitly with SCHOOL_DEBUG=1 on a trusted machine.
    debug = os.environ.get("SCHOOL_DEBUG") == "1"
    print(f"  Starting server at http://{host}:{port}  (debug={debug})")
    if host != "127.0.0.1" and not os.environ.get("SCHOOL_ROBOT_KEY"):
        print("  WARNING: listening beyond localhost with SCHOOL_ROBOT_KEY unset —")
        print("           /api/robot/* will refuse every request (fail closed),")
        print("           but /api/* has NO auth and exposes student names.")
    print("=" * 60 + "\n")
    app.run(debug=debug, host=host, port=port)
