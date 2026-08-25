"""
Robot-facing API for the school timetable app (AI-SHA integration).

Mounted as a Flask blueprint at /api/robot/*. Kept separate from app.py so the
timetable logic stays untouched and the security-sensitive surface is small
enough to read in one sitting.

WHY THIS EXISTS AS A SEPARATE SURFACE
The existing /api/* endpoints return full lists of named students. Those are
minors. AI-SHA stands in a public corridor with a speaker and a screen, so the
robot must never receive a list it could read out. This blueprint enforces the
split:

    SPEAKABLE  -> counts and class timetables. No individual is identified, so
                  the robot may receive and say these.
    EMAILED    -> any list of named students or teachers. Rendered and sent HERE;
                  the robot receives only {ok, count} and says "the list has been
                  emailed". Personal data never reaches the robot's memory, disk
                  or logs.

Every endpoint requires the shared secret in `x-robot-key` and FAILS CLOSED when
that secret is unset — an unconfigured server must never behave like an open one.

Configuration (all environment variables):
    SCHOOL_ROBOT_KEY       shared secret the robot presents      (required)
    SCHOOL_REPORT_TO       comma-separated recipients            (required to email)
    SCHOOL_SMTP_HOST/PORT  default smtp.gmail.com / 587
    SCHOOL_SMTP_USER       sending account
    SCHOOL_SMTP_PASS       app password, never the real password
    SCHOOL_BELL_JSON       path to a bell-schedule JSON (see BELL_SCHEDULE below)
"""

from __future__ import annotations

import hmac
import json
import os
import smtplib
from datetime import datetime, time as _time
from email.message import EmailMessage
from html import escape
from zoneinfo import ZoneInfo

import re
from flask import Blueprint, jsonify, request

robot_api = Blueprint("robot_api", __name__)

SCHOOL_TZ = ZoneInfo(os.environ.get("SCHOOL_TZ", "Asia/Dubai"))

# ── Bell schedule ────────────────────────────────────────────────────────────
# The timetable app has no concept of "now" — every endpoint takes an explicit
# day and period. Answering "who is free right now?" needs wall-clock -> period,
# which is school policy, not data we hold.
#
# ⚠️ THESE TIMES ARE PLACEHOLDERS. They are a plausible shape, NOT ISC Sharjah's
# real bell schedule. Point SCHOOL_BELL_JSON at the real one before anybody
# relies on a "right now" answer, or the robot will confidently report the wrong
# period. Format: {"Period 1": ["07:45", "08:30"], ...}
DEFAULT_BELL_SCHEDULE = {
    "Period 1": ["07:45", "08:30"],
    "Period 2": ["08:30", "09:15"],
    "Period 3": ["09:15", "10:00"],
    "Period 4": ["10:20", "11:05"],
    "Period 5": ["11:05", "11:50"],
    "Period 6": ["11:50", "12:35"],
    "Period 7": ["13:15", "14:00"],
    "Period 8": ["14:00", "14:45"],
}


def _load_bell_schedule() -> tuple[dict, bool]:
    """(schedule, is_real). is_real is False when we fell back to placeholders."""
    path = os.environ.get("SCHOOL_BELL_JSON")
    if path and os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f), True
        except (OSError, ValueError) as e:
            print(f"[robot_api] bell schedule at {path} unreadable ({e}); using placeholders")
    return DEFAULT_BELL_SCHEDULE, False


def _parse_hhmm(s: str) -> _time:
    h, m = s.split(":")
    return _time(int(h), int(m))


def resolve_now() -> dict:
    """Current school day and period, or the reason there isn't one."""
    schedule, is_real = _load_bell_schedule()
    now = datetime.now(SCHOOL_TZ)
    day = now.strftime("%A")
    current = None
    for name, (start, end) in schedule.items():
        if _parse_hhmm(start) <= now.time() <= _parse_hhmm(end):
            current = name
            break
    return {
        "day": day,
        "period": current,
        "time": now.strftime("%H:%M"),
        "in_session": current is not None,
        "bell_schedule_is_real": is_real,
    }


# ── Enrolment guard ──────────────────────────────────────────────────────────
# The app calls a student free if their section has no class OR the student is not
# enrolled in the subject being taught. That second rule reads a "Subject
# Enrollment" sheet. When that sheet is EMPTY every student looks un-enrolled in
# everything, so the app reports ALL of them free — a confident, catastrophic
# wrong answer for the person asking. Refuse those questions instead.
_ENROL_CACHE: dict = {}


def enrolment_present() -> bool:
    if "v" in _ENROL_CACHE:
        return _ENROL_CACHE["v"]
    ok = False
    try:
        import openpyxl
        from app import STUDENT_XLSX
        wb = openpyxl.load_workbook(STUDENT_XLSX, read_only=True, data_only=True)
        if "Subject Enrollment" in wb.sheetnames:
            ws = wb["Subject Enrollment"]
            ok = ws.max_row is not None and ws.max_row > 1   # more than the header
        wb.close()
    except Exception as e:                                    # noqa: BLE001
        print(f"[robot_api] could not check enrolment data: {e}")
    _ENROL_CACHE["v"] = ok
    return ok


def student_grades() -> list:
    """Grades that actually have students. The timetable covers 2-12 but the
    master list only 5-12, so iterating the timetable's grades would produce
    meaningless zero-rows for 2-4."""
    import openpyxl
    from app import STUDENT_XLSX
    out = set()
    wb = openpyxl.load_workbook(STUDENT_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = None
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(h).strip() if h else "" for h in row]
            continue
        g = dict(zip(hdr, row)).get("Class")
        if g not in (None, ""):
            try:
                out.add(int(str(g)))
            except ValueError:
                pass
    wb.close()
    return sorted(out)


# ── period label -> teacher-schedule slot ────────────────────────────────────
# The two data sources number time differently and this is not obvious from
# either file.  The class timetable NAMES periods ("Period 6", "SLO/Period 6");
# the teacher schedule NUMBERS every slot in the day INCLUDING lunch.  Lunch is
# the teacher schedule's slot 6 and nobody teaches in it, so that source has no
# slot "6" at all - the numbering agrees up to period 5 and is off by one after.
#
# Derived empirically, not guessed: for every (teacher, day) pair the subject in
# the class timetable was matched against each teacher-schedule slot, and the
# winning slot counted.  Period 6 -> 7 (68 matches), Period 7 -> 8 (81),
# Period 8 -> 9 (152), Period 9 -> 10 (139); periods 1-5 map to themselves.
#
# Passing "6" straight through matched nothing, so no teacher was ever marked
# busy, the endpoint reported 127 of 127 free, and busy_free_sane() refused the
# answer.  The guard was right; the lookup was wrong.
_TEACHER_SLOT = {1: "1", 2: "2", 3: "3", 4: "4", 5: "5",
                 6: "7", 7: "8", 8: "9", 9: "10"}


def teacher_slot(period_label) -> str:
    """Map a class-timetable period label to the teacher schedule's slot number.

    Accepts "Period 6", "SLO/Period 6", "SLO/P7", "6", "After School".
    Returns "" when the label carries no period (e.g. "Lunch"), which callers
    already treat as "not in session".
    """
    s = str(period_label or "").strip()
    if not s:
        return ""
    if "after school" in s.lower():
        return "11"
    m = re.search(r"(\d+)", s)
    if not m:
        return ""
    return _TEACHER_SLOT.get(int(m.group(1)), m.group(1))


def busy_free_sane(n_free: int, n_busy: int) -> bool:
    """Reject the "matched nothing" shape: everybody free, nobody busy.

    During a normal school period most teachers are teaching. free=127 busy=0 is
    not a quiet day, it is a lookup that compared the wrong things - and it is
    exactly what a period-format mismatch produces. Refusing beats emailing an
    administrator a list of every teacher in the school.
    """
    return not (n_busy == 0 and n_free > 5)


def audit_free(free: list, busy: list) -> dict:
    """Spot free-counts that are artefacts rather than facts.

    Two ways this app reports students free when they are not:

    1. A grade with no students in the master list (it covers 5-12, the timetable
       covers 2-12) yields total=0 and a meaningless "0 free".
    2. A section whose period is a COMBINED label - "Math/English/2nd Lang",
       "Math Alg / Geo" - can never match any single enrolment name, so every
       student in it is judged "not enrolled" and therefore free. Only ~2% of
       cells, but where it hits the answer is 100% wrong.

    A section in which EVERY student is free specifically because they are "not
    enrolled" is the signature of (2): a real free period reports "Free period".
    """
    total = len(free) + len(busy)
    if total == 0:
        return {"reliable": False, "reason": "no_students_for_grade"}
    per_sec: dict = {}
    for s in free:
        d = per_sec.setdefault(s.get("section"), {"free": 0, "busy": 0, "notenrol": 0})
        d["free"] += 1
        if str(s.get("reason", "")).startswith("Not enrolled"):
            d["notenrol"] += 1
    for s in busy:
        per_sec.setdefault(s.get("section"), {"free": 0, "busy": 0, "notenrol": 0})["busy"] += 1
    suspect = [sec for sec, d in per_sec.items()
               if d["busy"] == 0 and d["free"] > 0 and d["notenrol"] == d["free"]]
    return {"reliable": not suspect, "suspect_sections": sorted(map(str, suspect)),
            "reason": "unmatched_combined_subject" if suspect else None}


def free_students_reportable(free: list, busy: list, audit: dict) -> bool:
    """Is a free-student result worth stating at all?

    audit_free() already flags sections it could not attribute, and both callers
    used to append a caveat and report the number anyway. That is fine when a
    section or two is affected. It is not fine when the caveat swallows the
    answer: a grade-8 period-1 query returned "269 of 269 students are free"
    with a warning attached, during a period when the entire grade was in class.
    A number that wrong is not improved by a footnote - the listener remembers
    269, and the emailed version is a list of every child in the grade.

    This is the student-side twin of busy_free_sane(): refuse the shapes that
    mean "the lookup matched nothing", and let everything else through with the
    existing caveat.
    """
    total = len(free) + len(busy)
    if total == 0:
        return False
    # Nobody busy during a scheduled period: the subject never matched an
    # enrolment, so "free" here means "not found", not "not in class".
    if not busy and len(free) > 5:
        return False
    # Most of the grade sits in sections we could not attribute.
    suspect = set(audit.get("suspect_sections") or [])
    if suspect:
        n_suspect = sum(1 for x in free if str(x.get("section")) in suspect)
        if n_suspect * 2 >= total:
            return False
    return True


_UNRELIABLE_FREE = ("I cannot answer that reliably. For that period the timetable "
                    "splits those sections across parallel groups, so I cannot tell "
                    "which students are actually free. Please check with the school "
                    "office on +971 6 558 2211.")


_NO_ENROL = {
    "ok": False,
    "reason": "no_enrolment_data",
    "speakable": ("I cannot answer that yet. I know the timetable, but not which "
                  "subjects each student takes, so I cannot tell who is free."),
}


# ── Auth ─────────────────────────────────────────────────────────────────────
def _refuse(request) -> tuple | None:
    expected = os.environ.get("SCHOOL_ROBOT_KEY")
    if not expected:
        print("[robot_api] SCHOOL_ROBOT_KEY is not configured; refusing request.")
        return jsonify({"error": "Robot endpoint not configured"}), 503
    provided = request.headers.get("x-robot-key", "")
    # compare_digest is constant-time; == would leak the secret through timing.
    if not provided or not hmac.compare_digest(provided, expected):
        return jsonify({"error": "Unauthorized"}), 401
    return None


# ── Email ────────────────────────────────────────────────────────────────────
def _recipients() -> list[str]:
    return [a.strip() for a in os.environ.get("SCHOOL_REPORT_TO", "").split(",") if a.strip()]


def send_report(subject: str, html: str) -> tuple[bool, str]:
    to = _recipients()
    if not to:
        return False, "SCHOOL_REPORT_TO is not set"
    user = os.environ.get("SCHOOL_SMTP_USER", "")
    pw = os.environ.get("SCHOOL_SMTP_PASS", "")
    if not user or not pw:
        return False, "SCHOOL_SMTP_USER / SCHOOL_SMTP_PASS are not set"

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = user
    msg["To"] = ", ".join(to)
    msg.set_content("This report is HTML. See the attached formatted version.")
    msg.add_alternative(html, subtype="html")
    try:
        with smtplib.SMTP(os.environ.get("SCHOOL_SMTP_HOST", "smtp.gmail.com"),
                          int(os.environ.get("SCHOOL_SMTP_PORT", "587")), timeout=30) as s:
            s.starttls()
            s.login(user, pw)
            s.send_message(msg)
        return True, f"sent to {len(to)} recipient(s)"
    except Exception as e:                                    # noqa: BLE001
        return False, f"{type(e).__name__}: {e}"


_CELL = "padding:8px;border:1px solid #e2e8f0;"


def _render_people(title: str, subtitle: str, rows: list[dict], columns: list[tuple[str, str]]) -> str:
    head = "".join(f'<td style="{_CELL}font-weight:bold;">{escape(lbl)}</td>' for lbl, _ in columns)
    body = "".join(
        "<tr>" + "".join(
            f'<td style="{_CELL}">{escape(str(r.get(key, "") or ""))}</td>' for _, key in columns
        ) + "</tr>"
        for r in rows
    )
    empty = "" if rows else '<p style="color:#64748b;">Nobody matched.</p>'
    return f"""<div style="font-family:Arial,sans-serif;max-width:700px;margin:0 auto;">
      <h2 style="color:#2563eb;">{escape(title)}</h2>
      <p>{escape(subtitle)}</p>
      {empty}
      {'' if not rows else f'<table style="border-collapse:collapse;width:100%;margin:16px 0;"><tr style="background:#f1f5f9;">{head}</tr>{body}</table>'}
      <hr style="border:none;border-top:1px solid #e2e8f0;margin:24px 0;" />
      <p style="color:#64748b;font-size:12px;">
        ISC Sharjah — timetable system. Requested via AI-SHA. Contains personal data
        about students; do not forward.
      </p></div>"""


# ── Endpoints ────────────────────────────────────────────────────────────────
# Imported lazily inside handlers: app.py loads the Excel/DB on first use, and a
# module-level import here would create a circular import at startup.

@robot_api.route("/api/robot/now")
def r_now():
    if (bad := _refuse(request)):
        return bad
    return jsonify(resolve_now())


@robot_api.route("/api/robot/timetable")
def r_timetable():
    """SPEAKABLE — a section's day. Institutional, names nobody."""
    if (bad := _refuse(request)):
        return bad
    from app import get_class_timetable

    grade = request.args.get("grade", type=int)
    section = request.args.get("section", "A")
    day = request.args.get("day") or resolve_now()["day"]
    if not grade:
        return jsonify({"error": "grade is required"}), 400
    try:
        periods = get_class_timetable(grade, section, day)
    except Exception as e:                                    # noqa: BLE001
        return jsonify({"error": f"timetable lookup failed: {e}"}), 500

    # get_class_timetable returns period_name / is_free, not period / free.
    spoken = "; ".join(
        f"{p.get('period_name')}: "
        f"{'free' if p.get('effectively_free') or not p.get('subject') else p['subject']}"
        for p in periods
    ) or "no lessons found"
    return jsonify({
        "ok": True, "grade": grade, "section": section, "day": day,
        "periods": periods,
        "speakable": f"Grade {grade} section {section} on {day}. {spoken}.",
    })


@robot_api.route("/api/robot/free-count")
def r_free_count():
    """SPEAKABLE — how many are free. Counts only, no names leave this process."""
    if (bad := _refuse(request)):
        return bad
    if not enrolment_present():
        return jsonify(_NO_ENROL), 200
    from app import find_free_students

    grade = request.args.get("grade", type=int)
    now = resolve_now()
    day = request.args.get("day") or now["day"]
    period = request.args.get("period") or now["period"]
    if not period:
        return jsonify({"ok": False, "reason": "not_in_session",
                        "speakable": "There is no lesson running at the moment."}), 200

    # "How many TEACHERS are free?" used to fall through to the student counter
    # and answer with a student number - the right shape, the wrong people. A
    # count names nobody, so it is spoken rather than emailed.
    if request.args.get("subject") == "teachers":
        from app import find_free_teachers
        pnum = teacher_slot(period)
        try:
            free_t, busy_t = find_free_teachers(day, pnum)
        except Exception as e:                                # noqa: BLE001
            return jsonify({"error": f"teacher lookup failed: {e}"}), 500
        if not busy_free_sane(len(free_t), len(busy_t)):
            return jsonify({"ok": False, "reason": "implausible_result",
                            "speakable": ("I got an implausible answer from the "
                                          "timetable system, so I am not reporting "
                                          "it.")}), 200
        n = len(free_t)
        return jsonify({"ok": True, "subject": "teachers", "day": day, "period": period,
                        "free_count": n, "busy_count": len(busy_t),
                        "speakable": (f"{n} teacher{'' if n == 1 else 's'} "
                                      f"{'is' if n == 1 else 'are'} free in {period} "
                                      f"on {day}.")})

    # No grade named = the whole school. "Which students are available now?" is a
    # perfectly natural question for a director and used to be refused outright.
    if not grade:
        totals, suspect_any, dropped = [], [], []
        for g in student_grades():
            try:
                f, b = find_free_students(g, day, period)
            except Exception:
                continue
            a = audit_free(f, b)
            if a["reason"] == "no_students_for_grade":
                continue
            if not free_students_reportable(f, b, a):
                # Summing a grade whose lookup matched nothing inflates the
                # school-wide total by the size of that grade. Drop it and say so.
                dropped.append(str(g))
                continue
            totals.append((g, len(f), len(b)))
            if not a["reliable"]:
                suspect_any.append(str(g))
        n = sum(t[1] for t in totals)
        speak = (f"Across grades {totals[0][0]} to {totals[-1][0]}, {n} students are "
                 f"free in {period} on {day}." if totals else
                 "I have no student records to count.")
        if suspect_any:
            speak += (" Treat that as an over-estimate: in grade"
                      f"{'s' if len(suspect_any) > 1 else ''} {', '.join(suspect_any)} "
                      "the timetable lists a combined subject I cannot match to "
                      "individual enrolments.")
        if dropped:
            speak += (f" Grade{'s' if len(dropped) > 1 else ''} {', '.join(dropped)} "
                      f"{'are' if len(dropped) > 1 else 'is'} not included: for that "
                      "period the timetable splits those sections across parallel "
                      "groups, so I cannot tell who is free.")
        return jsonify({"ok": True, "scope": "all_grades", "day": day, "period": period,
                        "free_count": n, "per_grade": [
                            {"grade": g, "free": f, "busy": b} for g, f, b in totals],
                        "reliable": not suspect_any, "speakable": speak})
    try:
        free, busy = find_free_students(grade, day, period)
    except Exception as e:                                    # noqa: BLE001
        return jsonify({"error": f"lookup failed: {e}"}), 500

    n = len(free)
    audit = audit_free(free, busy)
    if audit["reason"] == "no_students_for_grade":
        return jsonify({"ok": False, "reason": "no_students_for_grade",
                        "speakable": (f"I have the timetable for grade {grade}, but no "
                                      "student records for it, so I cannot count who "
                                      "is free.")}), 200
    if not free_students_reportable(free, busy, audit):
        return jsonify({"ok": False, "reason": "unattributable_result",
                        "grade": grade, "day": day, "period": period,
                        "suspect_sections": audit.get("suspect_sections"),
                        "speakable": _UNRELIABLE_FREE}), 200
    speak = (f"In grade {grade}, {n} student{'' if n == 1 else 's'} "
             f"{'is' if n == 1 else 'are'} free in {period} on {day}.")
    if not audit["reliable"]:
        # Say so rather than quietly overstating. The administrator can act on a
        # caveat; they cannot act on a number that is silently wrong.
        speak += (" Treat that as an over-estimate: in section"
                  f"{'s' if len(audit['suspect_sections']) > 1 else ''} "
                  f"{', '.join(audit['suspect_sections'])} the timetable lists a "
                  "combined subject I cannot match to individual enrolments.")
    return jsonify({
        "ok": True, "grade": grade, "day": day, "period": period,
        "free_count": n, "busy_count": len(busy),
        "reliable": audit["reliable"], "suspect_sections": audit["suspect_sections"],
        "speakable": speak,
    })


@robot_api.route("/api/robot/free-students")
def r_free_students():
    """EMAILED — the named list. The robot gets counts only, never the names."""
    if (bad := _refuse(request)):
        return bad
    if not enrolment_present():
        return jsonify(_NO_ENROL), 200
    from app import find_free_students

    grade = request.args.get("grade", type=int)
    now = resolve_now()
    day = request.args.get("day") or now["day"]
    period = request.args.get("period") or now["period"]
    if not period:
        return jsonify({"ok": False, "reason": "not_in_session"}), 200

    if not grade:                     # whole school, one grouped report
        rows, suspect, dropped = [], [], []
        for g in student_grades():
            try:
                f, b = find_free_students(g, day, period)
            except Exception:
                continue
            if not f and not b:
                continue
            a = audit_free(f, b)
            if not free_students_reportable(f, b, a):
                # Do NOT put these names in the email. When the lookup matches
                # nothing every child in the grade is listed as "free", and the
                # recipient gets a roster of the whole grade presented as a
                # finding. Naming them is the harm; a caveat does not undo it.
                dropped.append(str(g))
                continue
            if not a["reliable"]:
                suspect.append(str(g))
            for s_ in f:
                s_ = dict(s_); s_["grade_label"] = str(g); rows.append(s_)
        # The COUNT endpoint warns when a grade's result is an artefact; the emailed
        # list must carry the same warning or the administrator acts on 1,204 names
        # of whom hundreds are actually sitting in class.
        note = ""
        if dropped:
            note += (f"<p style=\"background:#fee2e2;border-left:4px solid #dc2626;"
                     f"padding:10px;margin:12px 0;\"><b>Grade"
                     f"{'s' if len(dropped) > 1 else ''} {', '.join(dropped)} "
                     f"{'are' if len(dropped) > 1 else 'is'} omitted.</b> For this "
                     f"period the timetable splits those sections across parallel "
                     f"groups, so every student in them would have been listed as "
                     f"free. Those names are deliberately not included.</p>")
        if suspect:
            note = (f"<p style=\"background:#fef3c7;border-left:4px solid #d97706;"
                    f"padding:10px;margin:12px 0;\"><b>Treat this list as an "
                    f"over-estimate.</b> In grade{'s' if len(suspect) > 1 else ''} "
                    f"{', '.join(suspect)} the timetable lists a combined subject "
                    f"(for example \"Math/English/2nd Lang\") that cannot be matched "
                    f"to individual subject enrolments, so every student in those "
                    f"sections is reported free. Grades not listed here are "
                    f"unaffected.</p>")
        html = _render_people(
            f"Free students — all grades, {period}, {day}",
            f"{len(rows)} student(s) free across grades "
            f"{student_grades()[0]}-{student_grades()[-1]}.",
            rows,
            [("Grade", "grade_label"), ("First name", "first_name"),
             ("Last name", "last_name"), ("Section", "section"),
             ("Computer no.", "computer_number"), ("Why free", "reason")])
        html = html.replace("<hr", note + "<hr", 1)
        ok, info = send_report(f"Free students — all grades, {day} {period}", html)
        if not ok:
            return jsonify({"ok": False, "count": len(rows), "emailed": False,
                            "error": info}), 502
        speak = "The list has been sent to the administrator's email."
        if suspect:
            speak += (f" It is flagged as an over-estimate for grade"
                      f"{'s' if len(suspect) > 1 else ''} {', '.join(suspect)}.")
        return jsonify({"ok": True, "scope": "all_grades", "count": len(rows),
                        "emailed": True, "recipients": len(_recipients()),
                        "reliable": not suspect, "suspect_grades": suspect,
                        "speakable": speak})
    try:
        free, _busy = find_free_students(grade, day, period)
    except Exception as e:                                    # noqa: BLE001
        return jsonify({"error": f"lookup failed: {e}"}), 500

    html = _render_people(
        f"Free students — Grade {grade}, {period}, {day}",
        f"{len(free)} student(s) free.",
        free,
        [("First name", "first_name"), ("Last name", "last_name"),
         ("Section", "section"), ("Computer no.", "computer_number"),
         ("Why free", "reason")],
    )
    ok, info = send_report(f"Free students — Grade {grade}, {day} {period}", html)
    if not ok:
        return jsonify({"ok": False, "count": len(free), "emailed": False, "error": info}), 502
    return jsonify({"ok": True, "count": len(free), "emailed": True,
                    "recipients": len(_recipients())})


@robot_api.route("/api/robot/free-teachers")
def r_free_teachers():
    """EMAILED — staff names are personal data too."""
    if (bad := _refuse(request)):
        return bad
    from app import find_free_teachers

    now = resolve_now()
    day = request.args.get("day") or now["day"]
    # find_free_teachers compares against the BARE period number from the teacher
    # schedule ("3"), while the rest of the app - and the robot's parser - speak in
    # "Period 3". Passing the long form matched nothing, so no teacher was ever
    # marked busy and the endpoint cheerfully reported 127 of 128 teachers free.
    # Normalise whatever we are given.
    period_raw = request.args.get("period") or now["period"] or ""
    period = teacher_slot(period_raw)
    if not period:
        return jsonify({"ok": False, "reason": "not_in_session"}), 200
    try:
        free, _busy = find_free_teachers(day, period)
    except Exception as e:                                    # noqa: BLE001
        return jsonify({"error": f"lookup failed: {e}"}), 500

    rows = free
    if not busy_free_sane(len(free), len(_busy)):
        return jsonify({"ok": False, "reason": "implausible_result",
                        "free": len(free), "busy": len(_busy),
                        "speakable": ("I got an implausible answer from the timetable "
                                      "system — it says nearly every teacher is free, "
                                      "which usually means the period did not match. "
                                      "I have not sent anything.")}), 200
    html = _render_people(
        f"Free teachers — period {period}, {day}",
        f"{len(rows)} teacher(s) free.",
        rows,
        [("Name", "name"), ("Code", "teacher_code"), ("Subject", "subject")],
    )
    ok, info = send_report(f"Free teachers — {day} period {period}", html)
    if not ok:
        return jsonify({"ok": False, "count": len(rows), "emailed": False, "error": info}), 502
    return jsonify({"ok": True, "count": len(rows), "emailed": True,
                    "recipients": len(_recipients())})
