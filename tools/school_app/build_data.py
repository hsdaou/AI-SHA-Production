#!/usr/bin/env python3
"""
Build the timetable app's data files from the school's own PDF reports.

    python3 build_data.py --classes "Classes G2 to 12.pdf" \
                          --teachers "Teachers G3 to 12.pdf" \
                          --roster Book2.xlsx --out .

Produces, in --out:
    school_timetable.db             class_sections + timetable_entries
    Teacher_Schedule_Database.xlsx  teacher directory + full schedule
    Student_Database_1.xlsx         student roster (+ Subject Enrollment sheet)
    bell_schedule.json              the REAL period times, read out of the PDF

WHY WORD COORDINATES, NOT `pdftotext -layout`
These are Crystal Reports timetable grids. A single period cell routinely holds
several wrapped lines ("-AAK1, Arabic Language- Arabs, 02A" over four lines), and
`-layout` interleaves neighbouring columns once that happens, silently welding
one subject onto another. `-bbox-layout` gives every word an x/y box, so cells can
be reconstructed geometrically: columns from the header positions, rows from the
day labels. Slower to write, but it does not quietly invent a timetable.

⚠️ ENROLMENT. The app decides a student is free if their section has no class OR
the student is not enrolled in the subject being taught. That second rule needs a
"Subject Enrollment" sheet mapping student -> subjects. A plain roster does not
contain it, and with the sheet EMPTY the app treats every student as not-enrolled
and reports ALL of them free. So this script writes an `enrollment_present` flag
into the roster workbook; robot_api refuses student-level answers when it is
false, rather than confidently reporting 2,717 free students.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
import subprocess
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# Teacher codes look like AAK1, EGO1, ETC2, SRB2 — 2-4 letters then a digit.
TEACHER_CODE_RE = re.compile(r"\b([A-Z]{2,4}\d)\b")
# Group headers look like "Gr AA2A-H" / "Gr PE2BEJ" / "Gr G2-Rl1".
GROUP_RE = re.compile(r"\bGr\s+([A-Za-z0-9\-]+)")
SECTION_RE = re.compile(r"Class-Section:\s*(\d{2}-[A-Z0-9]+)")


# ── PDF -> words ─────────────────────────────────────────────────────────────
def pdf_words(path: str):
    """Yield (page_index, [ {text,x0,x1,y0,y1} ]) using pdftotext -bbox-layout."""
    xml = subprocess.run(
        ["pdftotext", "-bbox-layout", path, "-"],
        capture_output=True, text=True, check=True).stdout
    # Crystal emits raw '&' in a few subject names, which breaks the XML parser.
    xml = re.sub(r"&(?!(amp|lt|gt|quot|apos);)", "&amp;", xml)
    root = ET.fromstring(xml)
    ns = {"x": "http://www.w3.org/1999/xhtml"}
    pages = root.findall(".//x:page", ns) or root.findall(".//page")
    for idx, page in enumerate(pages):
        words = []
        for w in (page.findall(".//x:word", ns) or page.findall(".//word")):
            if w.text and w.text.strip():
                words.append({
                    "text": w.text.strip(),
                    "x0": float(w.get("xMin")), "x1": float(w.get("xMax")),
                    "y0": float(w.get("yMin")), "y1": float(w.get("yMax")),
                })
        yield idx, words


def group_lines(words, tol=3.0):
    """Cluster words into visual lines by y, then sort each line by x."""
    lines = []
    for w in sorted(words, key=lambda w: (w["y0"], w["x0"])):
        for ln in lines:
            if abs(ln["y"] - w["y0"]) <= tol:
                ln["words"].append(w)
                break
        else:
            lines.append({"y": w["y0"], "words": [w]})
    for ln in lines:
        ln["words"].sort(key=lambda w: w["x0"])
        ln["text"] = " ".join(w["text"] for w in ln["words"])
    return sorted(lines, key=lambda l: l["y"])


# ── Class timetable pages ────────────────────────────────────────────────────
def parse_class_page(words):
    """One 'Class-Section: NN-X' page -> (section_code, period_defs, cells)."""
    lines = group_lines(words)

    section = None
    for ln in lines:
        m = SECTION_RE.search(ln["text"])
        if m:
            section = m.group(1)
            break
    if not section:
        return None

    # Header row: the line containing "Period 1". Its words define the columns.
    hdr = next((ln for ln in lines if re.search(r"\bPeriod\s+1\b", ln["text"])), None)
    if not hdr:
        return None

    # Column labels can be multi-word ("SLO/Period 6", "After School"); merge
    # words that sit within a few points of each other into one heading.
    cols = []
    for w in hdr["words"]:
        if cols and w["x0"] - cols[-1]["x1"] < 6:
            cols[-1]["label"] += " " + w["text"]
            cols[-1]["x1"] = w["x1"]
        else:
            cols.append({"label": w["text"], "x0": w["x0"], "x1": w["x1"]})

    # "After School" appears TWICE (3:20-4:10 and 4:10-5:00). Keyed by label alone
    # the second silently overwrites the first and both cells weld together.
    seen = {}
    for c in cols:
        n = seen.get(c["label"], 0) + 1
        seen[c["label"]] = n
        if n > 1:
            c["label"] = f'{c["label"]} ({n})'

    # Column x-boundaries: midpoints between adjacent header centres. Computed
    # BEFORE the times are read, because the times row must be split on exactly
    # the same boundaries — a per-header tolerance lets "8:50AM" (the END of
    # Period 1) drift into Period 2's column and shifts every period by one.
    centres = [(c["x0"] + c["x1"]) / 2 for c in cols]
    bounds = []
    for i in range(len(cols)):
        lo = 0 if i == 0 else (centres[i - 1] + centres[i]) / 2
        hi = 10_000 if i == len(cols) - 1 else (centres[i] + centres[i + 1]) / 2
        bounds.append((lo, hi))

    # Times row: the first line below the header that looks like times.
    times, times_bottom = {}, hdr["y"]
    for ln in lines:
        if ln["y"] > hdr["y"] and re.search(r"\d{1,2}:\d{2}", ln["text"]):
            for w in ln["words"]:
                cx = (w["x0"] + w["x1"]) / 2
                ci = next((i for i, (lo, hi) in enumerate(bounds) if lo <= cx < hi), None)
                if ci is not None:
                    times[cols[ci]["label"]] = times.get(cols[ci]["label"], "") + w["text"]
            times_bottom = max(w["y1"] for w in ln["words"])
            break

    # Day labels give the row bands.
    day_rows = [(ln["y"], ln["words"][0]["text"])
                for ln in lines
                if ln["words"] and ln["words"][0]["text"] in DAYS]
    if not day_rows:
        return None
    day_rows.sort()

    # Row y-bands: a day's band runs from just above its label to the next day.
    band = []
    for i, (y, day) in enumerate(day_rows):
        # Clamp to below the times row. Without this the first day's band reaches
        # up into it and every Monday subject comes out as "8:00AM- 8:50AM English",
        # which also makes empty cells look non-empty and Lunch look like a lesson.
        top = max(y - 22, times_bottom + 1) if i == 0 else (day_rows[i - 1][0] + y) / 2
        bottom = y + 60 if i == len(day_rows) - 1 else (y + day_rows[i + 1][0]) / 2
        band.append((day, top, bottom))

    day_label_x = max(w["x1"] for ln in lines for w in ln["words"]
                      if w["text"] in DAYS) if day_rows else 0

    cells = defaultdict(list)
    for w in words:
        if w["text"] in DAYS or w["x1"] <= day_label_x:
            continue                                  # the day-label gutter
        cy = (w["y0"] + w["y1"]) / 2
        cx = (w["x0"] + w["x1"]) / 2
        day = next((d for d, t, b in band if t <= cy < b), None)
        if day is None:
            continue
        ci = next((i for i, (lo, hi) in enumerate(bounds) if lo <= cx < hi), None)
        if ci is None:
            continue
        cells[(day, cols[ci]["label"])].append(w)

    period_defs = [(c["label"], times.get(c["label"], "")) for c in cols]
    return section, period_defs, cells


def cell_to_entry(cell_words):
    """Turn a cell's words into (subject, teacher_code, group_code, is_free, raw)."""
    if not cell_words:
        return None, None, None, 1, ""
    raw = " ".join(w["text"] for w in group_lines(cell_words)[i]["words"]
                   for i in [0]) if False else " ".join(
        ln["text"] for ln in group_lines(cell_words))
    raw = re.sub(r"\s+", " ", raw).strip()

    if not raw or raw == "//":
        return None, None, None, 1, raw

    group = None
    m = GROUP_RE.search(raw)
    if m:
        group = m.group(1)

    codes = TEACHER_CODE_RE.findall(raw)
    teacher = codes[0] if codes else None

    # Subject = the text with the group header, teacher codes and section lists
    # stripped out. What survives is the lesson name.
    s = GROUP_RE.sub(" ", raw)
    s = re.sub(r"-?\b[A-Z]{2,4}\d\b,?", " ", s)          # teacher codes
    s = re.sub(r"\b\d{2}[A-Z]\b,?", " ", s)              # section refs like 02A
    s = re.sub(r"\bK\d[A-Z]\b,?", " ", s)                # kindergarten refs
    s = s.replace("//", " ")
    s = re.sub(r"[,\-]{2,}", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" ,-")
    # One cell can list the same subject once per teacher ("Arabic Language- Arabs,
    # Arabic Language- Arabs, ..."). Keep the distinct phrases in order.
    parts, seen_p = [], set()
    for p in (p.strip(" -") for p in s.split(",")):
        key = p.lower()
        if p and key not in seen_p:
            seen_p.add(key)
            parts.append(p)
    s = ", ".join(parts)
    subject = s or None
    return subject, teacher, group, 0 if subject else 1, raw


def build_timetable_db(classes_pdf, out_db):
    if os.path.exists(out_db):
        os.remove(out_db)
    conn = sqlite3.connect(out_db)
    conn.executescript("""
        CREATE TABLE class_sections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            grade INTEGER NOT NULL,
            section_code TEXT NOT NULL UNIQUE
        );
        CREATE TABLE timetable_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_section_id INTEGER NOT NULL REFERENCES class_sections(id),
            day TEXT NOT NULL,
            period_name TEXT NOT NULL,
            period_time TEXT,
            subject TEXT,
            teacher_code TEXT,
            group_code TEXT,
            is_free INTEGER NOT NULL DEFAULT 0,
            raw_cell TEXT
        );
        CREATE INDEX idx_te_lookup ON timetable_entries(class_section_id, day);
    """)

    bell, n_sections, n_entries, skipped = {}, 0, 0, 0
    for page_idx, words in pdf_words(classes_pdf):
        parsed = parse_class_page(words)
        if not parsed:
            skipped += 1
            continue
        section_code, period_defs, cells = parsed
        grade = int(section_code.split("-")[0])
        cur = conn.execute(
            "INSERT OR IGNORE INTO class_sections (grade, section_code) VALUES (?,?)",
            (grade, section_code))
        sid = conn.execute("SELECT id FROM class_sections WHERE section_code=?",
                           (section_code,)).fetchone()[0]
        if cur.rowcount:
            n_sections += 1

        for label, time_str in period_defs:
            if time_str and label not in bell:
                bell[label] = time_str

        for (day, period), cw in cells.items():
            subject, teacher, group, is_free, raw = cell_to_entry(cw)
            conn.execute(
                "INSERT INTO timetable_entries (class_section_id, day, period_name,"
                " period_time, subject, teacher_code, group_code, is_free, raw_cell)"
                " VALUES (?,?,?,?,?,?,?,?,?)",
                (sid, day, period, dict(period_defs).get(period, ""),
                 subject, teacher, group, is_free, raw))
            n_entries += 1

    conn.commit()
    conn.close()
    return n_sections, n_entries, skipped, bell


# ── Teacher schedule pages ───────────────────────────────────────────────────
def parse_teacher_page(words):
    lines = group_lines(words)
    code = name = None
    for ln in lines:
        m = re.search(r"Teacher Code:\s*([A-Z0-9]+)", ln["text"])
        if m:
            code = m.group(1)
        m = re.search(r"Teacher Name:\s*(.+)$", ln["text"])
        if m:
            name = m.group(1).strip()
        if code and name:
            break
    if not code:
        return None

    hdr = next((ln for ln in lines
                if re.fullmatch(r"(\d+\s+){3,}\d+", ln["text"].strip())), None)
    if not hdr:
        return code, name, []

    cols = [{"label": w["text"], "x0": w["x0"], "x1": w["x1"]} for w in hdr["words"]]
    centres = [(c["x0"] + c["x1"]) / 2 for c in cols]
    bounds = []
    for i in range(len(cols)):
        lo = 0 if i == 0 else (centres[i - 1] + centres[i]) / 2
        hi = 10_000 if i == len(cols) - 1 else (centres[i] + centres[i + 1]) / 2
        bounds.append((lo, hi))

    day_rows = sorted((ln["y"], ln["words"][0]["text"]) for ln in lines
                      if ln["words"] and ln["words"][0]["text"] in DAYS)
    if not day_rows:
        return code, name, []
    band = []
    for i, (y, day) in enumerate(day_rows):
        top = y - 18 if i == 0 else (day_rows[i - 1][0] + y) / 2
        bottom = y + 60 if i == len(day_rows) - 1 else (y + day_rows[i + 1][0]) / 2
        band.append((day, top, bottom))
    label_x = max(w["x1"] for ln in lines for w in ln["words"] if w["text"] in DAYS)

    cells = defaultdict(list)
    for w in words:
        if w["text"] in DAYS or w["x1"] <= label_x:
            continue
        cy, cx = (w["y0"] + w["y1"]) / 2, (w["x0"] + w["x1"]) / 2
        day = next((d for d, t, b in band if t <= cy < b), None)
        ci = next((i for i, (lo, hi) in enumerate(bounds) if lo <= cx < hi), None)
        if day and ci is not None:
            cells[(day, cols[ci]["label"])].append(w)

    rows = []
    for (day, period), cw in cells.items():
        raw = re.sub(r"\s+", " ", " ".join(ln["text"] for ln in group_lines(cw))).strip()
        if not raw:
            continue
        secs = re.findall(r"\b(?:\d{2}[A-Z]|K\d[A-Z])\b", raw)
        subject = re.sub(r"\(.*?\)", " ", raw)
        subject = re.sub(r"\b(?:\d{2}[A-Z]|K\d[A-Z])\b,?", " ", subject)
        subject = re.sub(r"\(?\s*\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}\s*\)?", " ", subject)
        subject = re.sub(r"\s+", " ", subject).strip(" ,-")
        rows.append({"day": day, "period": period, "subject": subject,
                     "sections": ",".join(dict.fromkeys(secs)), "raw": raw})
    return code, name, rows


def build_teacher_xlsx(teachers_pdf, out_xlsx):
    import openpyxl
    wb = openpyxl.Workbook()
    d = wb.active
    d.title = "Teacher Directory"
    d.append(["Teacher Code", "Title", "Teacher Name", "Total Periods/Week",
              "Subjects Taught", "Rooms/Classes"])
    sch = wb.create_sheet("Teacher Schedule")
    sch.append(["Teacher Code", "Day", "Period", "Subject", "Sections", "Raw"])

    n_t = n_r = 0
    for _idx, words in pdf_words(teachers_pdf):
        parsed = parse_teacher_page(words)
        if not parsed:
            continue
        code, name, rows = parsed
        title = ""
        if name:
            m = re.match(r"(Mr\.|Ms\.|Mrs\.|Dr\.)\s*(.*)", name)
            if m:
                title, name = m.group(1), m.group(2)
        subjects = sorted({r["subject"] for r in rows if r["subject"]})
        d.append([code, title, name, len(rows), "; ".join(subjects), ""])
        n_t += 1
        for r in rows:
            sch.append([code, r["day"], r["period"], r["subject"], r["sections"], r["raw"]])
            n_r += 1
    wb.save(out_xlsx)
    return n_t, n_r


# ── Student roster ───────────────────────────────────────────────────────────
def build_student_xlsx(roster_xlsx, out_xlsx):
    """Roster -> the two-sheet workbook the app expects.

    Returns (n_students, enrollment_present). The Subject Enrollment sheet is
    created with headers so the app does not KeyError, but it is EMPTY unless the
    source actually carried enrolment rows — and that emptiness is recorded, not
    hidden, because an empty sheet makes the app call every student free.
    """
    import openpyxl
    src = openpyxl.load_workbook(roster_xlsx, read_only=True, data_only=True)
    ws = src[src.sheetnames[0]]

    out = openpyxl.Workbook()
    m = out.active
    m.title = "Students"
    m.append(["Computer Number", "First Name", "Last Name", "Class", "Section",
              "Email", "Number of Subjects"])

    headers, n = None, 0
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else "" for h in row]
            continue
        rec = dict(zip(headers, row))
        num = rec.get("Computer Number") or rec.get("Student Number")
        if num is None:
            continue
        m.append([str(num), rec.get("First Name", ""), rec.get("Last Name", ""),
                  rec.get("Class", ""), rec.get("Section", ""),
                  rec.get("Email", ""), rec.get("Number of Subjects", 0)])
        n += 1
    src.close()

    enr = out.create_sheet("Subject Enrollment")
    enr.append(["Class", "Section", "Computer Number", "Subject Code",
                "Subject Name", "Teacher", "Course Group"])

    meta = out.create_sheet("_meta")
    meta.append(["key", "value"])
    meta.append(["enrollment_present", "false"])
    meta.append(["note", "Subject Enrollment is EMPTY. The app treats a student with "
                         "no enrolments as NOT enrolled in whatever is being taught, "
                         "i.e. FREE — so student-level free/busy answers would be "
                         "wrong for everyone. robot_api refuses them while this is false."])
    out.save(out_xlsx)
    return n, False


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--classes", required=True)
    ap.add_argument("--teachers", required=True)
    ap.add_argument("--roster")
    ap.add_argument("--out", default=".")
    a = ap.parse_args()
    os.makedirs(a.out, exist_ok=True)

    print("[1/4] class timetables ...")
    db = os.path.join(a.out, "school_timetable.db")
    ns, ne, skipped, bell = build_timetable_db(a.classes, db)
    print(f"      {ns} sections, {ne} timetable cells  ({skipped} pages had no grid)")

    print("[2/4] bell schedule ...")
    bell_out = {}
    for label, t in bell.items():
        if not re.match(r"^(Period|SLO|Lunch|After School)", label):
            continue                      # stray column artefacts, not real periods
        m = re.findall(r"(\d{1,2}):(\d{2})\s*([AP]M)", t)
        if len(m) >= 2:
            def to24(h, mi, ap):
                h = int(h) % 12 + (12 if ap == "PM" else 0)
                return f"{h:02d}:{mi}"
            bell_out[label] = [to24(*m[0]), to24(*m[1])]
    bj = os.path.join(a.out, "bell_schedule.json")
    json.dump(bell_out, open(bj, "w"), indent=2)
    print(f"      {len(bell_out)} periods -> {os.path.basename(bj)}")
    for k, v in bell_out.items():
        print(f"        {k:<16} {v[0]}-{v[1]}")

    print("[3/4] teacher schedules ...")
    tx = os.path.join(a.out, "Teacher_Schedule_Database.xlsx")
    nt, nr = build_teacher_xlsx(a.teachers, tx)
    print(f"      {nt} teachers, {nr} scheduled periods")

    if a.roster:
        print("[4/4] student roster ...")
        sx = os.path.join(a.out, "Student_Database_1.xlsx")
        nstu, enr = build_student_xlsx(a.roster, sx)
        print(f"      {nstu} students; enrolment data present: {enr}")
        if not enr:
            print("      !! WITHOUT enrolment, student-level 'who is free' cannot be")
            print("         answered correctly — every student would look free.")
    else:
        print("[4/4] student roster ... skipped (no --roster)")

    print("\ndone.")


if __name__ == "__main__":
    sys.exit(main())
